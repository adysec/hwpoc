#!/usr/bin/env python3
"""Convert hw漏洞 Excel to individual TOML files organized by year."""

import openpyxl
import re
import unicodedata
from pathlib import Path

SRC = Path("/home/adysec/图片/poc/hw漏洞（真）by AdySec.xlsx")
DST = Path(__file__).resolve().parent.parent / "docs" / "_data"

HEADER_MAP = {
    "2025": ["vendor", "name", "type", "version", "path", "detail", "fix", "date", "label", "verifier"],
    "2024": ["vendor", "name", "type", "version", "path", "detail", "fix", "date", "label"],
    "2023": ["vendor", "name", "type", "source", "info", "date", "poc", "label"],
}

def clean(val):
    if val is None:
        return ""
    val = str(val).strip()
    val = re.sub(r'\s+', ' ', val)
    return val

def slug(s):
    s = re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff\u3400-\u4dbf_-]', '-', s)
    s = re.sub(r'-+', '-', s).strip('-').lower()
    return s[:100] or "untitled"

def toml_val(v):
    v = v.replace('\\', '\\\\').replace('"', '\\"')
    return f'"{v}"'

def toml_entry(keys, vals):
    lines = []
    for k, v in zip(keys, vals):
        if v:
            lines.append(f'{k} = {toml_val(v)}')
    return '\n'.join(lines) + '\n'

def main():
    wb = openpyxl.load_workbook(SRC)

    for year, cols in HEADER_MAP.items():
        if year not in wb.sheetnames:
            continue

        ws = wb[year]
        year_dir = DST / year
        existing = set(year_dir.iterdir()) if year_dir.exists() else set()
        if existing:
            for f in existing:
                f.unlink()

        count = 0
        for r in range(2, ws.max_row + 1):
            row_vals = [clean(ws.cell(r, c).value) for c in range(1, len(cols) + 1)]
            if not any(row_vals):
                continue

            vendor_slug = slug(row_vals[0]) if row_vals[0] else "unknown"
            name_slug = slug(row_vals[1]) if len(row_vals) > 1 and row_vals[1] else "unknown"
            fname = f"{vendor_slug}_{name_slug}.toml"

            content = toml_entry(cols, row_vals)
            (year_dir / fname).write_text(content, encoding="utf-8")
            count += 1

        print(f"{year}: {count} files → {year_dir}")

if __name__ == "__main__":
    main()
